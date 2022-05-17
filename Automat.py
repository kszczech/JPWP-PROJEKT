import openpyxl as opx
import openpyxl.styles as styles
import openpyxl.utils as ou

def fill(column_in, column_out, inputDir, outputDir, fileDir, fromInput, toInput, fromOutput, toOutput):
    redFill = styles.PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    input_file = opx.load_workbook(inputDir)
    output_file = opx.load_workbook(outputDir)
    wsInput = input_file.active
    wsOutput = output_file.active
    input_tab = []
    output_tab = []
    column_date = []
    #column_in =  ['B','C','D','E','F','G','H','I','J','L',"M","N"] #wpisz tu z inputa kolumny do przetworzenia
    #column_out = ['C','D','E','F','T','U','V','G','H','X',"Y","Z"] #wpisz tu z outputu kolumny ktore odpowiadaja kolumna z inputu
    for i in range(len(column_in)):
        line = []
        line.append(column_in[i])
        line.append(column_out[i])
        column_date.append(line)

    #print(column_date)
    #print("pytanie do inputu")
    y1 = fromInput #int(input("od ktorego wiersza zaczac"))
    y2 = toInput #int(input("do ktorego wiersza czytac"))

    for row in range(y1, y2+1):
        line = []
        for col in column_date:
            char = col[0] + str(row)
            line.append(wsInput[char].value)
        input_tab.append(line)
    #print(input_tab)

    print()
    #print("pytania do outputu")
    y1 = fromOutput#int(input("od ktorego wiersza zaczac"))
    y2 = toOutput#int(input("do ktorego wiersza czytac"))

    for row in range(y1, y2+1):
        line = []
        for col in column_date:
            char = col[1] + str(row)
            line.append(wsOutput[char].value)
        output_tab.append(line)
    #print(output_tab)

    #print(output_tab[1][0])

    for out in range(len(output_tab)):
        for inp in range(len(input_tab)):
            if output_tab[out][0] == input_tab[inp][0] and output_tab[out][1] == input_tab[inp][1]:
                for i in range(len(output_tab[out])):
                    if output_tab[out][i] == None:
                        output_tab[out][i] = input_tab[inp][i]
                        char = "A" + str(inp + 2)
                        wsInput[char].fill = redFill


#print(output_tab)

    for i in range(len(output_tab)):
        for j in range(len(column_date)):
            char = column_date[j][1] + str(i + y1)
            wsOutput[char].value = output_tab[i][j]

    nazwa = fileDir#input("jak nazwac gotowy plik")
    output_file.save(nazwa + '.xlsx')
    input_file.save(nazwa + "_input"+".xlsx")


def rooms():
    ist_people = []
    #excelowa część

    hotel = []
    list_room = [[2, 3], [3, 2], [4, 1], [5, 2], [6, 3]] #[wielkosc pokoju, ilosc takich pokoi]

    for room in list_room:
        for i in range(room[1]):
            line = []
            for j in range(room[0]):
                line.append("")
            line.append(room[0])
            hotel.append(line)

    print(hotel)

    
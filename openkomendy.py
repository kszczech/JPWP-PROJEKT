workbookDirectory = "przykładowy.xlsx"
from openpyxl.utils import get_column_letter
import openpyxl as opx

people = [
    ["Jan", "Kowalski", "123456789", "401234", "posilek miesny"],
    ["Beniamin", "Nowak", "987654321", "391234", "posilek wegetariański"],
    ["Alicja", "Kowalska", "908765432", "411234", "posiłek wegański"],
]

wb = opx.load_workbook(workbookDirectory)

worksheet = wb.active
worksheet.title = "Nowy tytuł"

worksheet1 = wb.create_sheet("Moj nowy worksheet")

col = 0
row = 0
for row in range(len(people)):
    for col in range(len(people[row])):
        char = get_column_letter(col+1) + str(row+1)
        worksheet[char].value = people[row][col]

wb.save(workbookDirectory)
